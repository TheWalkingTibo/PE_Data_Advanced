import pandas as pandas
import matplotlib.pyplot as plot

from PEDataAdvanced.classes.categorizer import Categorizer
from PEDataAdvanced.classes.date_generator import DateGenerator

import numpy as np
import statistics

from openpyxl import load_workbook

plot.rcdefaults()
row_count = 100
football_excel_file = pandas.read_excel("voetbal.xlsx")

print(football_excel_file)

# --- Opgave 2: Genereren van datums ---

print("\n")

randomDates = []
for i in range(row_count):
    randomDate = DateGenerator.generate_date_between("1/1/2011", "1/1/2012")
    randomDates.append(randomDate)

print(randomDates) #weergave datum



dataframe = pandas.DataFrame({'geboortedatum': randomDates})

book = load_workbook('voetbal.xlsx')
writer = pandas.ExcelWriter('voetbal.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

dataframe.to_excel(writer, 'gegevens', startrow=1, startcol= 3, header=False, index=False)

writer.save()



# --- Opgave 3: Categorizeren van maanden ---

engagement_categories = []
for i in range(row_count):
    engagement = Categorizer.categorize_date(randomDates[i])
    engagement_categories.append(engagement)

print(engagement_categories)


dataframe = pandas.DataFrame({'inzet': engagement_categories})

book = load_workbook('voetbal.xlsx')
writer = pandas.ExcelWriter('voetbal.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

dataframe.to_excel(writer, 'gegevens', startrow=1, startcol= 4, header=False, index=False)

writer.save()



# --- Opgave 4: Grafiek genereren ---


def get_column_from_sheet(column_name):
    column = []
    for row in football_excel_file.index:
        column.append(football_excel_file[column_name][row])
    return column


weights = get_column_from_sheet("gewicht")
lengths = get_column_from_sheet("lengte")

plot.scatter(weights, lengths)
plot.show()

# --- Opgave 5: Staafdiagram ---

positions = get_column_from_sheet("positie")
goals = get_column_from_sheet("aantal gemaakte goalen")

goalsByPosition = {
    "keeper": 0,
    "staart": 0,
    "linkervleugel": 0,
    "rechtervleugel": 0,
    "piloot": 0,
}

for i in range(row_count):
    position = positions[i]
    goalsByPosition[position] = goalsByPosition[position] + goals[i]

print("rechterflank doelpunten" + str(goalsByPosition['keeper']))

onderkant = np.arange(len(goalsByPosition.keys()))
grafiekhoogte = [v for v in goalsByPosition.values()]

plot.bar(onderkant, grafiekhoogte, align='center', alpha=0.5)
plot.xticks(onderkant, goalsByPosition.keys())
plot.ylabel('gescoorde doelpunten')
plot.show()



effort = {
    "zeer goed": 0,
    "goed": 0,
    "matig": 0
}

for i in range(row_count):
    j = engagement_categories[i]
    effort[j] = effort[j] + 1

onderkant = np.arange(len(effort.keys()))
grafiekhoogte = [v for v in effort.values()]

plot.bar(onderkant, grafiekhoogte, align='center', alpha=0.5)
plot.xticks(onderkant, effort.keys())
plot.ylabel('inzet')
plot.show()



### Opgave 6: Modus en gemiddelde kolom D

def calculate_average():
    geheel = 0
    for i in range(row_count):
        geheel = geheel + goals[i]

    return geheel / row_count


print(calculate_average())
print(max(goals))


# --- Opgave 7: Kwartiel 1 en standaarddeviatie berekenen

eersteKwartiel = np.percentile(weights, 25) # kwartiel 1 berekenen

print("Eerste kwartiel: " + str(eersteKwartiel))

standaardafwijking = statistics.stdev(weights) # standaardafwijking
print("Standaardafwijking: " + str(standaardafwijking))

# --- Opgave 8: verband tussen posities en aantal goals

print("ja, hoe dieper op het veld, hoe meer goalen dat je maakt, kijk naar de staafdiagram, daaraan zie je dat keepers, geen goalen scoren, terwijl piloten veel doelpunten scoren")
#indien nodig vergelijken met met de 2 arrays (resultaat piloten scoren meer dan staart en deze op hun beurt dan keepers)

# --- Opgave 9:

categories = {
    "zeer goed": 0,
    "goed": 0,
    "matig": 0
}
for i in range(row_count):
    categorie = engagement_categories[i]
    categories[categorie] = categories[categorie] + 1

labels = 'zeer goed', 'goed', 'matig'
sizes = [categories["zeer goed"],categories["goed"],categories["matig"]]
colors = ['blue', 'red', 'green']

plot.pie(sizes, labels = labels, colors = colors)
plot.axis('equal')
plot.show()


# --- Opgave 10:

goalsByPosition = {
    "linkervleugel": [],
    "rechtervleugel": [],
    "piloot": []
}

for i in range(row_count):
    position = positions[i]
    if position in goalsByPosition:
        goalsByPosition[position].append(goals[i])

for goals in goalsByPosition.values():
    plot.boxplot(goals)
    plot.show()

# --- Opgave 11: soorten gegevens

print("kwantitatief, discreet")       #aantal gemaakte goalen
print("kwalitatief, ordinaal")       #inzet
print("kwantitatief, continue")       #gewicht